# -*- coding: utf-8 -*-
# @Time    : 2022/3/2 11:07 AM 
# @Author  : Dou Chen
# @File    : pythonTest2.py

import os  # 系统库
import openpyxl as opxl
import re  # 正则库

# 快速创建 名称为: 学号姓名-1.py 文件，不用手动创建（累）
def quicklyCreatePyFile(fileNameList):
    folderPath = '../作业/'
    for item in fileNameList:
        filePath = folderPath + item  # 组织文件名
        newFile = open(filePath, 'w')  # 'w' 模式在没有该文件的情况下会新建
        newFile.close()  # 关闭文件
    fileList = os.listdir(folderPath)  # 检查路径下的文件，查看是否创建成功
    print(fileList)  # 输出文件夹下的文件
    print('creating finish !')

# 所有交了作业的学生的 学号 和 姓名，用于创建 作业 目录下的.py文件
def getFileNameList(filePath):
    fileNamList = []  # 存放文件名
    wb = opxl.load_workbook(filePath)
    sheet = wb.worksheets[0]
    rows = sheet.max_row  # 获取行数
    for i in range(2, rows + 1):
        fileName = sheet.cell(i, 1).value + sheet.cell(i, 2).value + '-1.py'
        fileNamList.append(fileName)
    # 额外的一个同学
    newFileName = '2021206190088邵子轩-1.py'
    fileNamList.append(newFileName)
    return fileNamList

# 获取output excel文件，包括用户的添加、成绩填写...
def getOutput(rPath, wPath):
    # 文件读取和创建工作
    r_wb = opxl.load_workbook(rPath)
    r_ws = r_wb.worksheets[0]
    rows = r_ws.max_row  # 最后一行
    cols = r_ws.max_column  # 最大列数
    w_wb = opxl.Workbook()  # 创建 workbook
    w_ws = w_wb.create_sheet('python课程平时成绩', 0)  # 创建表
    existStudent = {}  # 记录已存在的学生信息，用于后续处理
    for i in range(1, rows + 1):
        for j in range(1, cols + 1):
            w_ws.cell(i, j).value = r_ws.cell(i, j).value  # 复制数据
        if i == 1:
            continue # 第一行是标题
        existStudent[str(w_ws.cell(i, 1).value)] = i  # 加入存在的用户字典，包括 学号 在第几行，用于后续查询
    folderPath = '../作业/'
    fileList = os.listdir(folderPath)  # 获取路径下的所有文件
    for file in fileList:
        if file.endswith('.py'):  # 判读是否为py文件
            numberPattern = re.compile(r'\d+')  # 正则表达式匹配学号
            stringPattern = re.compile(r'[\u4e00-\u9fa5]+')  # 正则表达式匹配姓名
            id = re.search(numberPattern, file).group()  # 获取学生学号
            name = re.findall(stringPattern, file)[0]  # 获取学生姓名
            if id in existStudent.keys():  # 如果原先成绩表中存在该学生
                w_ws.cell(existStudent[id], 3).value = os.stat(rPath).st_mtime  # 写入 rPath 最后修改时间为 平时成绩1
                w_ws.cell(existStudent[id], 4).value = os.stat(wPath).st_mtime  # 写入 wPath 最后修改时间为 平时成绩2
            else: # 反之创建新的一行 填入信息
                w_ws.cell(rows + 1, 1).value = id  # 写入学号
                w_ws.cell(rows + 1, 2).value = name  # 写入姓名
                w_ws.cell(rows + 1, 3).value = os.stat(rPath).st_mtime  # 写入 rPath 最后修改时间为 平时成绩1
                w_ws.cell(rows + 1, 4).value = os.stat(wPath).st_mtime  # 写入 wPath 最后修改时间为 平时成绩2
                rows += 1  # 最大行数 + 1
    print('writing finish !')
    w_wb.save(wPath)
    w_wb.close()

if __name__ == '__main__':

    # ----- 此部分用于 自动创建作业python文件 ----- （准备工作）
    filePath = './成绩表.xlsx'
    fileNameList = getFileNameList(filePath)  # 获取需要创建的py文件名
    quicklyCreatePyFile(fileNameList)  # 在 作业 目录下自动创建所有py文件

    # ----- 此部分用于 读写任务 ----- （作业内容）
    rPath = './成绩表.xlsx'
    wPath = '../out_data/output.xlsx'
    getOutput(rPath, wPath)


